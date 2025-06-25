import openpyxl
from openpyxl import Workbook

class InventoryManager:
    def __init__(self, file_name="warehouse.xlsx"):
        self.file_name = file_name
        self.workbook = self._load_or_create_workbook()
        self.sheet = self.workbook.active

    def _load_or_create_workbook(self):
        """Загружает существующий файл Excel или создаёт новый"""
        try:
            workbook = openpyxl.load_workbook(self.file_name)
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Inventory"
            # Добавляем заголовки с категорией
            sheet.append(["ID", "Название", "Количество", "Категория"])
            workbook.save(self.file_name)
        return workbook

    def update_quantity(self, item_name, change):
        """Обновляет количество предмета в инвентаре"""
        for row in self.sheet.iter_rows(min_row=2, values_only=False):
            if row[1].value == item_name:  # Сравниваем название предмета
                row[2].value += change  # Обновляем количество
                self.workbook.save(self.file_name)
                return
        raise ValueError(f"Предмет с названием '{item_name}' не найден в инвентаре.")

    def get_quantity(self, item_name):
        """Возвращает количество предмета по его названию"""
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == item_name:  # Сравниваем название предмета
                return row[2]  # Возвращаем количество
        return 0  # Если предмет не найден, возвращаем 0

    def get_inventory(self):
        """Возвращает список предметов из инвентаря"""
        inventory = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            # row[3] — это Категория
            inventory.append({
                "ID": row[0],
                "Название": row[1],
                "Количество": row[2],
                "Категория": row[3] if len(row) > 3 else None
            })
        return inventory